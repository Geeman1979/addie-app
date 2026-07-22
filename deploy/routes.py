from flask import render_template, request, redirect, url_for, flash, jsonify, abort
from flask_login import login_user, logout_user, login_required, current_user
from datetime import datetime
import json

def register_routes(app, db):

    from models import User, Request, Phase, Approval, KirkpatrickResult, InterviewGuide, ContentTag
    from models import LNACycle, TrainingEvent, ResourceAllocation, DepartmentBudget
    from models import LearningStatistic, LearningUploadBatch, Cohort, AuditLog, ResourceForecast
    from models import PhaseActivity
    from datetime import datetime, date, timedelta
    import csv, io

    def learning_expert_required(f):
        """Decorator: only learning experts can perform write actions"""
        from functools import wraps
        @wraps(f)
        def decorated(*args, **kwargs):
            if not current_user.is_authenticated or not current_user.is_learning_expert():
                flash('Only learning experts can perform this action', 'danger')
                return redirect(url_for('request_list'))
            return f(*args, **kwargs)
        return decorated

    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if request.method == 'POST':
            user = User.query.filter_by(email=request.form['email']).first()
            if user and user.check_password(request.form['password']):
                login_user(user)
                return redirect(url_for('dashboard'))
            flash('Invalid email or password', 'danger')
        return render_template('login.html')

    @app.route('/logout')
    @login_required
    def logout():
        logout_user()
        return redirect(url_for('login'))

    @app.route('/change-password', methods=['GET', 'POST'])
    @login_required
    def change_password():
        if request.method == 'POST':
            if not current_user.check_password(request.form['current_password']):
                flash('Current password is incorrect', 'danger')
                return redirect(url_for('change_password'))
            if request.form['new_password'] != request.form['confirm_password']:
                flash('New passwords do not match', 'danger')
                return redirect(url_for('change_password'))
            if len(request.form['new_password']) < 6:
                flash('Password must be at least 6 characters', 'danger')
                return redirect(url_for('change_password'))
            current_user.set_password(request.form['new_password'])
            db.session.commit()
            logout_user()
            flash('Password changed. Please login with your new password.', 'success')
            return redirect(url_for('login'))
        return render_template('change_password.html')

    @app.route('/')
    @login_required
    def dashboard():
        total = Request.query.count()
        in_progress = Request.query.filter(Request.status.in_(['In Analysis','In Design','In Development','In Implementation','In Evaluation'])).count()
        completed = Request.query.filter_by(status='Completed').count()
        new_requests = Request.query.filter_by(status='New').count()
        critical = Request.query.filter_by(priority='Critical').count()
        escalation = Request.query.filter(Request.escalation_level>0).count()

        phase_counts = {}
        for p in ['inbox','analysis','design','development','implementation','evaluation']:
            phase_counts[p] = Request.query.filter_by(current_phase=p).count()

        recent = Request.query.order_by(Request.created_at.desc()).limit(8).all()

        kirk_summary = db.session.query(
            KirkpatrickResult.level,
            db.func.avg(KirkpatrickResult.metric_value).label('avg_score'),
            db.func.count(KirkpatrickResult.id).label('count')
        ).group_by(KirkpatrickResult.level).all()

        blocker_count = 0
        for ph in Phase.query.filter(Phase.status=='In Progress').all():
            data = json.loads(ph.data) if ph.data else {}
            if data.get('blocker'): blocker_count += 1

        # Kirkpatrick projected vs actual
        proj_kirk = {'l1': '', 'l2': '', 'l3': '', 'l4': ''}
        act_kirk = {'l1': '', 'l2': '', 'l3': '', 'l4': ''}
        for p in Phase.query.filter_by(phase_name='analysis').all():
            d = json.loads(p.data) if p.data else {}
            for k in ['projected_l1', 'projected_l2', 'projected_l3', 'projected_l4']:
                if d.get(k):
                    short = k.replace('projected_', '')
                    if proj_kirk[short]: proj_kirk[short] += ' / '
                    proj_kirk[short] += str(d[k])
        for p in Phase.query.filter_by(phase_name='evaluation').all():
            d = json.loads(p.data) if p.data else {}
            for k in ['actual_l1', 'actual_l2', 'actual_l3', 'actual_l4']:
                if d.get(k):
                    short = k.replace('actual_', '')
                    if act_kirk[short]: act_kirk[short] += ' / '
                    act_kirk[short] += str(d[k])

        return render_template('dashboard.html', total=total, in_progress=in_progress,
            completed=completed, new_requests=new_requests, critical=critical,
            escalation=escalation, phase_counts=phase_counts, recent=recent,
            kirk_summary=kirk_summary, blocker_count=blocker_count,
            proj_kirk=proj_kirk, act_kirk=act_kirk, datetime=datetime)

    @app.route('/requests')
    @login_required
    def request_list():
        status_filter = request.args.get('status', 'all')
        priority_filter = request.args.get('priority', 'all')
        search_query = request.args.get('search', '').strip()
        query = Request.query
        if status_filter != 'all': query = query.filter_by(status=status_filter)
        if priority_filter != 'all': query = query.filter_by(priority=priority_filter)
        if search_query:
            q = f'%{search_query}%'
            query = query.filter(
                db.or_(Request.title.ilike(q), Request.requester_name.ilike(q),
                       Request.requester_department.ilike(q), Request.description.ilike(q))
            )
        requests = query.order_by(Request.created_at.desc()).all()
        priority_order = {'Critical': 0, 'High': 1, 'Medium': 2, 'Low': 3, 'Backlog': 4}
        requests.sort(key=lambda r: priority_order.get(r.priority, 5))
        users = User.query.order_by(User.name).all()
        return render_template('requests.html', requests=requests, status_filter=status_filter,
            priority_filter=priority_filter, search_query=search_query, users=users)

    @app.route('/requests/new', methods=['GET', 'POST'])
    @login_required
    def new_request():
        if request.method == 'POST':
            req = Request(title=request.form['title'], description=request.form['description'],
                priority=request.form['priority'], request_type=request.form.get('request_type','planned'),
                lna_cycle_id=int(request.form['lna_cycle_id']) if request.form.get('lna_cycle_id') else None, requester_name=request.form['requester_name'],
                requester_department=request.form['requester_department'],
                business_need=request.form['business_need'],
                expected_outcome=request.form['expected_outcome'],
                target_audience=request.form['target_audience'],
                estimated_learners=int(request.form['estimated_learners'] or 0),
                deadline=datetime.strptime(request.form['deadline'], '%Y-%m-%d').date() if request.form['deadline'] else None,
                is_adhoc=request.form.get('is_adhoc') == 'on',
                projected_roi=float(request.form['projected_roi'] or 0),
                assigned_to_id=int(request.form['assigned_to_id']) if request.form.get('assigned_to_id') else None,
                status='Pending Review', review_status='Pending', current_phase='inbox')
            db.session.add(req)
            db.session.commit()
            log_audit('Created request', 'Request', req.id, f'{req.title}')
            flash('Request created successfully', 'success')
            return redirect(url_for('request_list'))
        users = User.query.all()
        cycles = LNACycle.query.filter_by(status='Planning').all()
        return render_template('new_request.html', users=users, cycles=cycles)

    @app.route('/requests/<int:req_id>/delete', methods=['POST'])
    @login_required
    def delete_request(req_id):
        """Delete a request/intervention and all associated data"""
        req = Request.query.get_or_404(req_id)
        title = req.title
        # Manually delete child records to avoid NOT NULL FK constraint errors
        Phase.query.filter_by(request_id=req_id).delete()
        Approval.query.filter_by(request_id=req_id).delete()
        KirkpatrickResult.query.filter_by(request_id=req_id).delete()
        InterviewGuide.query.filter_by(request_id=req_id).delete()
        ContentTag.query.filter_by(request_id=req_id).delete()
        ResourceAllocation.query.filter_by(request_id=req_id).delete()
        LearningStatistic.query.filter_by(request_id=req_id).delete()
        LearningUploadBatch.query.filter_by(request_id=req_id).delete()
        log_audit('Deleted request', 'Request', req_id, f'{title}')
        db.session.delete(req)
        db.session.commit()
        flash(f'Request "{title}" deleted successfully', 'success')
        return redirect(url_for('request_list'))

    @app.route('/requests/<int:req_id>/edit', methods=['POST'])
    @login_required
    def edit_request_title(req_id):
        """Edit the title of a request/intervention"""
        req = Request.query.get_or_404(req_id)
        new_title = request.form.get('title', '').strip()
        if new_title:
            old_title = req.title
            req.title = new_title
            db.session.commit()
            log_audit('Edited request title', 'Request', req_id, f'"{old_title}" → "{new_title}"')
            flash(f'Intervention renamed to "{new_title}"', 'success')
        return redirect(url_for('request_list'))

    @app.route('/requests/<int:req_id>/assign', methods=['POST'])
    @login_required
    def assign_request(req_id):
        req = Request.query.get_or_404(req_id)
        user_id = request.form.get('assigned_to_id', '').strip()
        if user_id:
            req.assigned_to_id = int(user_id)
            assigned_user = User.query.get(int(user_id))
            log_audit('Assigned request', 'Request', req_id, f'Assigned to {assigned_user.name if assigned_user else "Unknown"}')
        else:
            req.assigned_to_id = None
            log_audit('Unassigned request', 'Request', req_id, 'Removed assignment')
        db.session.commit()
        flash('Assignment updated', 'success')
        return redirect(url_for('request_list'))

    @app.route('/requests/<int:req_id>/review', methods=['POST'])
    @login_required
    def review_request(req_id):
        if not current_user.is_learning_expert():
            flash('Only learning experts can review', 'danger')
            return redirect(url_for('request_list'))
        req = Request.query.get_or_404(req_id)
        action = request.form.get('review_action', '')
        comment = request.form.get('review_comment', '')
        if action == 'Accepted':
            req.review_status = 'Accepted'
            req.review_comment = comment
            req.status = 'Accepted'
            # Auto-start the inbox phase when accepted
            phase = Phase.query.filter_by(request_id=req_id, phase_name='inbox').first()
            if not phase:
                phase = Phase(request_id=req_id, phase_name='inbox', status='In Progress', data='{}', started_at=datetime.utcnow())
                db.session.add(phase)
            else:
                phase.status = 'In Progress'
                phase.started_at = datetime.utcnow()
            req.current_phase = 'inbox'
            db.session.commit()
            log_audit('Request accepted', 'Request', req_id, f'Accepted. Note: {comment[:200]}')
            flash('Request accepted — inbox phase auto-started', 'success')
        elif action == 'Rejected':
            req.review_status = 'Rejected'
            req.review_comment = comment
            req.status = 'Rejected'
            for phase in req.phases.all():
                phase.status = 'Cancelled'
            req.current_phase = 'inbox'
            db.session.commit()
            log_audit('Request rejected', 'Request', req_id, f'Rejected. Note: {comment[:200]}')
            flash('Request rejected', 'warning')
        return redirect(url_for('request_list'))

    @app.route('/requests/<int:req_id>')
    @login_required
    def request_detail(req_id):
        req = Request.query.get_or_404(req_id)
        phases = {p.phase_name: p for p in req.phases.all()}
        approvals = req.approvals.order_by(Approval.created_at.desc()).all()
        interviews = InterviewGuide.query.filter_by(request_id=req_id).all()
        kirk = KirkpatrickResult.query.filter_by(request_id=req_id).order_by(KirkpatrickResult.level).all()
        tags = ContentTag.query.filter_by(request_id=req_id).all()
        users = User.query.all()
        batches = LearningUploadBatch.query.filter_by(request_id=req_id).order_by(LearningUploadBatch.uploaded_at.desc()).all()
        cohorts_list = Cohort.query.filter_by(request_id=req_id).all()
        activities = PhaseActivity.query.filter_by(request_id=req_id).order_by(PhaseActivity.start_date).all()
        return render_template('request_detail.html', req=req, phases=phases,
            approvals=approvals, interviews=interviews, kirk=kirk, tags=tags, users=users,
            batches=batches, cohorts_list=cohorts_list, activities=activities, json=json)

    @app.route('/requests/<int:req_id>/export/<fmt>')
    @login_required
    def export_request(req_id, fmt):
        """Export a single intervention to Excel"""
        req = Request.query.get_or_404(req_id)
        phases = {p.phase_name: p for p in req.phases.all()}
        approvals = req.approvals.order_by(Approval.created_at.desc()).all()
        stats = req.learning_stats.all()
        cohorts = Cohort.query.filter_by(request_id=req_id).all()

        from openpyxl import Workbook
        from openpyxl.styles import Font as XlFont, PatternFill as XlFill
        from openpyxl.utils import get_column_letter as get_column_letter

        wb = Workbook()

        # Sheet 1: Overview
        ws = wb.active; ws.title = 'Overview'
        ws.cell(row=1, column=1, value=f'Intervention: {req.title}').font = XlFont(bold=True, size=14)
        headers = ['Field', 'Value']
        for c, h in enumerate(headers, 1):
            ws.cell(row=2, column=c, value=h).font = XlFont(bold=True)
        data_rows = [('Status', req.status), ('Priority', req.priority), ('Current Phase', req.current_phase),
            ('Type', req.request_type), ('Requester', req.requester_name or ''),
            ('Department', req.requester_department or ''), ('Target Audience', req.target_audience or ''),
            ('Estimated Learners', req.estimated_learners or 0), ('Deadline', str(req.deadline or '')),
            ('Created', str(req.created_at.date() if req.created_at else '')),
            ('Assigned To', req.assigned_to.name if req.assigned_to else 'Unassigned'),
            ('Business Need', req.business_need or ''), ('Expected Outcome', req.expected_outcome or ''),
            ('Projected ROI', f'R{req.projected_roi:,.0f}' if req.projected_roi else '')]
        for ri, (k, v) in enumerate(data_rows, 3):
            ws.cell(row=ri, column=1, value=k); ws.cell(row=ri, column=2, value=str(v))
        ws.column_dimensions['A'].width = 25; ws.column_dimensions['B'].width = 55

        # Sheet 2: Learning Stats
        ws2 = wb.create_sheet('Learning Stats')
        stat_headers = ['Learner', 'Email', 'Department', 'Status', 'Completion %', 'Score', 'Minutes', 'Passed', 'Cohort']
        for c, h in enumerate(stat_headers, 1):
            ws2.cell(row=1, column=c, value=h).font = XlFont(bold=True)
        for ri, s in enumerate(stats, 2):
            ws2.cell(row=ri, column=1, value=s.learner_name or '')
            ws2.cell(row=ri, column=2, value=s.learner_email or '')
            ws2.cell(row=ri, column=3, value=s.department or '')
            ws2.cell(row=ri, column=4, value=s.completion_status or '')
            ws2.cell(row=ri, column=5, value=s.completion_pct or 0)
            ws2.cell(row=ri, column=6, value=s.score or 0)
            ws2.cell(row=ri, column=7, value=s.time_spent_minutes or 0)
            ws2.cell(row=ri, column=8, value='Yes' if s.passed else ('No' if s.passed is False else ''))
            ws2.cell(row=ri, column=9, value=s.cohort.name if s.cohort else '')
        for c in range(1, 10): ws2.column_dimensions[get_column_letter(c)].width = 20

        # Sheet 3: Approvals
        ws3 = wb.create_sheet('Approvals')
        app_headers = ['Phase', 'Person', 'Status', 'Comments', 'Date']
        for c, h in enumerate(app_headers, 1):
            ws3.cell(row=1, column=c, value=h).font = XlFont(bold=True)
        for ri, a in enumerate(approvals, 2):
            ws3.cell(row=ri, column=1, value=a.phase_name)
            ws3.cell(row=ri, column=2, value=a.approver_name or (a.approver.name if a.approver else ''))
            ws3.cell(row=ri, column=3, value=a.status)
            ws3.cell(row=ri, column=4, value=a.comments or '')
            ws3.cell(row=ri, column=5, value=str(a.decided_at or ''))
        for c in range(1, 6): ws3.column_dimensions[get_column_letter(c)].width = 25

        # Sheet 4: Phases
        ws4 = wb.create_sheet('ADDIE Phases')
        phase_headers = ['Phase', 'Status', 'Key Data']
        for c, h in enumerate(phase_headers, 1):
            ws4.cell(row=1, column=c, value=h).font = XlFont(bold=True)
        ri = 2
        for pname in ['inbox', 'analysis', 'design', 'development', 'implementation', 'evaluation']:
            p = phases.get(pname)
            pd = json.loads(p.data) if p and p.data else {}
            ws4.cell(row=ri, column=1, value=pname.capitalize())
            ws4.cell(row=ri, column=2, value=p.status if p else 'Not Started')
            # Condense phase data into readable text
            summary = '; '.join(f'{k}: {v}' for k, v in pd.items() if v) if pd else ''
            ws4.cell(row=ri, column=3, value=summary[:500])
            ri += 1
        ws4.column_dimensions['A'].width = 20; ws4.column_dimensions['B'].width = 15
        ws4.column_dimensions['C'].width = 80

        # Sheet 5: Cohorts
        if cohorts:
            ws5 = wb.create_sheet('Cohorts')
            coh_headers = ['Name', 'Description', 'Learners', 'Status', 'Start', 'End']
            for c, h in enumerate(coh_headers, 1):
                ws5.cell(row=1, column=c, value=h).font = XlFont(bold=True)
            for ri, coh in enumerate(cohorts, 2):
                ws5.cell(row=ri, column=1, value=coh.name)
                ws5.cell(row=ri, column=2, value=coh.description or '')
                ws5.cell(row=ri, column=3, value=coh.learner_count or 0)
                ws5.cell(row=ri, column=4, value=coh.status)
                ws5.cell(row=ri, column=5, value=str(coh.start_date or ''))
                ws5.cell(row=ri, column=6, value=str(coh.end_date or ''))
            for c in range(1, 7): ws5.column_dimensions[get_column_letter(c)].width = 22

        import io as io_mod
        from flask import make_response
        output = io_mod.BytesIO()
        wb.save(output)
        output.seek(0)
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename=intervention_{req_id}_{req.title[:30].replace(" ","_")}.xlsx'
        response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response

    @app.route('/requests/<int:req_id>/phase/<phase_name>/start', methods=['POST'])
    @login_required
    def start_phase(req_id, phase_name):
        if not current_user.is_learning_expert():
            flash('Only learning experts can manage phases', 'danger')
            return redirect(url_for('request_detail', req_id=req_id))
        req = Request.query.get_or_404(req_id)
        phase = Phase.query.filter_by(request_id=req_id, phase_name=phase_name).first()
        if not phase:
            phase = Phase(request_id=req_id, phase_name=phase_name, status='In Progress', data='{}', started_at=datetime.utcnow())
            db.session.add(phase)
        else:
            phase.status = 'In Progress'
            phase.started_at = datetime.utcnow()
        req.current_phase = phase_name
        req.status = f'In {phase_name.capitalize()}'
        db.session.commit()
        flash(f'{phase_name.capitalize()} phase started', 'success')
        return redirect(url_for('request_detail', req_id=req_id))

    @app.route('/requests/<int:req_id>/phase/<phase_name>/revert', methods=['POST'])
    @login_required
    def revert_phase(req_id, phase_name):
        """Revert a completed phase back to In Progress"""
        req = Request.query.get_or_404(req_id)
        phase = Phase.query.filter_by(request_id=req_id, phase_name=phase_name).first()
        if phase:
            phase.status = 'In Progress'
            phase.completed_at = None
            db.session.commit()
            log_audit('Reverted phase', 'Phase', req_id, f'{phase_name} reverted to In Progress')
            flash(f'{phase_name.capitalize()} phase reverted to In Progress', 'warning')
        return redirect(url_for('request_detail', req_id=req_id))

    @app.route('/requests/<int:req_id>/phase/<phase_name>/activities/add', methods=['POST'])
    @login_required
    def add_phase_activity(req_id, phase_name):
        """Add an activity to a phase"""
        a = PhaseActivity(
            request_id=req_id,
            phase_name=phase_name,
            activity_name=request.form['activity_name'],
            description=request.form.get('description', ''),
            owner_id=int(request.form['owner_id']) if request.form.get('owner_id') else None,
            start_date=datetime.strptime(request.form['start_date'], '%Y-%m-%d').date(),
            end_date=datetime.strptime(request.form['end_date'], '%Y-%m-%d').date(),
            status=request.form.get('status', 'In Progress')
        )
        db.session.add(a); db.session.commit()
        log_audit('Added phase activity', 'PhaseActivity', a.id, f'{a.activity_name} in {phase_name}')
        flash(f'Activity "{a.activity_name}" added', 'success')
        return redirect(url_for('request_detail', req_id=req_id))

    @app.route('/activities/<int:aid>/delete', methods=['POST'])
    @login_required
    def delete_phase_activity(aid):
        a = PhaseActivity.query.get_or_404(aid); req_id = a.request_id
        db.session.delete(a); db.session.commit()
        flash('Activity deleted', 'success')
        return redirect(url_for('request_detail', req_id=req_id))

    @app.route('/requests/<int:req_id>/phase/<phase_name>/save', methods=['POST'])
    @login_required
    def save_phase(req_id, phase_name):
        req = Request.query.get_or_404(req_id)
        phase = Phase.query.filter_by(request_id=req_id, phase_name=phase_name).first()
        data = json.loads(phase.data) if phase and phase.data else {}
        # Handle both single and multi-value fields
        seen_keys = set()
        for key in request.form:
            if key in ['csrf_token']:
                continue
            seen_keys.add(key)
            vals = request.form.getlist(key)
            if len(vals) > 1:
                data[key] = ','.join(v for v in vals if v)
            elif len(vals) == 1:
                data[key] = vals[0]
            else:
                data[key] = ''
        # Handle checkboxes that weren't checked (present in form but with empty values)
        multi_keys = ['data_sources', 'content_tags', 'keywords']
        for mk in multi_keys:
            if mk not in seen_keys and mk not in data:
                data[mk] = ''
        if not phase:
            phase = Phase(request_id=req_id, phase_name=phase_name, status='In Progress', data=json.dumps(data), started_at=datetime.utcnow())
            db.session.add(phase)
        else:
            phase.data = json.dumps(data)
            phase.updated_at = datetime.utcnow()
        # Sync the Request-level projected_roi from the form field
        if 'projected_roi_update' in data:
            try:
                req.projected_roi = float(data['projected_roi_update']) if data['projected_roi_update'] else None
            except ValueError:
                req.projected_roi = None
        # Only update current_phase if this phase is the one being actively worked on
        # (don't move backwards if user is saving data on a completed phase)
        if not phase or phase.status != 'Completed':
            req.current_phase = phase_name
        db.session.commit()
        flash(f'{phase_name.capitalize()} phase updated', 'success')
        return redirect(url_for('request_detail', req_id=req_id))

    @app.route('/requests/<int:req_id>/phase/<phase_name>/complete', methods=['POST'])
    @login_required
    def complete_phase(req_id, phase_name):
        phase = Phase.query.filter_by(request_id=req_id, phase_name=phase_name).first()
        if phase:
            phase.status = 'Completed'
            phase.completed_at = datetime.utcnow()
        next_phases = ['inbox','analysis','design','development','implementation','evaluation']
        idx = next_phases.index(phase_name)
        next_p = next_phases[idx + 1] if idx + 1 < len(next_phases) else None
        req = Request.query.get_or_404(req_id)
        if next_p:
            req.current_phase = next_p
            req.status = f'In {next_p.capitalize()}'
        else:
            req.current_phase = 'completed'
            req.status = 'Completed'
        db.session.commit()
        flash(f'{phase_name.capitalize()} phase completed', 'success')
        return redirect(url_for('request_detail', req_id=req_id))

    @app.route('/requests/<int:req_id>/phase/<phase_name>/blocker', methods=['POST'])
    @login_required
    def set_blocker(req_id, phase_name):
        phase = Phase.query.filter_by(request_id=req_id, phase_name=phase_name).first()
        data = json.loads(phase.data) if phase and phase.data else {}
        data['blocker'] = request.form['blocker']
        if phase:
            phase.data = json.dumps(data)
            phase.status = 'Blocked'
        db.session.commit()
        flash('Blocker logged', 'warning')
        return redirect(url_for('request_detail', req_id=req_id))

    @app.route('/requests/<int:req_id>/approve', methods=['POST'])
    @login_required
    def add_approval(req_id):
        approval = Approval(request_id=req_id, phase_name=request.form['phase_name'],
            approver_name=request.form.get('approver_name',''),
            status=request.form.get('status','Not Started'))
        db.session.add(approval)
        db.session.commit()
        log_audit('Added approval', 'Approval', approval.id, f'{approval.phase_name} — {approval.approver_name}')
        flash('Approval added', 'success')
        return redirect(url_for('request_detail', req_id=req_id))

    @app.route('/approvals/<int:approval_id>/respond', methods=['POST'])
    @login_required
    def respond_approval(approval_id):
        approval = Approval.query.get_or_404(approval_id)
        approval.status = request.form['status']
        approval.comments = request.form.get('comments', '')
        approval.decided_at = datetime.utcnow()
        db.session.commit()
        flash(f'Approval {approval.status.lower()}', 'success')
        return redirect(url_for('dashboard'))

    @app.route('/requests/<int:req_id>/escalate', methods=['POST'])
    @login_required
    def escalate_request(req_id):
        req = Request.query.get_or_404(req_id)
        req.escalation_level = (req.escalation_level or 0) + 1
        req.escalation_note = request.form['escalation_note']
        db.session.commit()
        flash(f'Escalated to level {req.escalation_level}', 'warning')
        return redirect(url_for('request_detail', req_id=req_id))

    @app.route('/requests/<int:req_id>/save-links', methods=['POST'])
    @login_required
    def save_links(req_id):
        """Save review link and LMS course link (always accessible from sidebar)"""
        req = Request.query.get_or_404(req_id)
        # Save review_link into development phase data
        dev_phase = Phase.query.filter_by(request_id=req_id, phase_name='development').first()
        dev_data = json.loads(dev_phase.data) if dev_phase and dev_phase.data else {}
        dev_data['review_link'] = request.form.get('review_link', '')
        if dev_phase:
            dev_phase.data = json.dumps(dev_data)
        else:
            dev_phase = Phase(request_id=req_id, phase_name='development', status='Not Started',
                data=json.dumps(dev_data), started_at=datetime.utcnow())
            db.session.add(dev_phase)

        # Save lms_course_link into implementation phase data
        impl_phase = Phase.query.filter_by(request_id=req_id, phase_name='implementation').first()
        impl_data = json.loads(impl_phase.data) if impl_phase and impl_phase.data else {}
        impl_data['lms_course_link'] = request.form.get('lms_course_link', '')
        if impl_phase:
            impl_phase.data = json.dumps(impl_data)
        else:
            impl_phase = Phase(request_id=req_id, phase_name='implementation', status='Not Started',
                data=json.dumps(impl_data), started_at=datetime.utcnow())
            db.session.add(impl_phase)

        db.session.commit()
        flash('Course links saved', 'success')
        return redirect(url_for('request_detail', req_id=req_id))

    @app.route('/requests/<int:req_id>/save-kirkpatrick', methods=['POST'])
    @login_required
    def save_kirkpatrick(req_id):
        """Save actual Kirkpatrick results (always accessible from sidebar, stored in evaluation phase data)"""
        req = Request.query.get_or_404(req_id)
        eval_phase = Phase.query.filter_by(request_id=req_id, phase_name='evaluation').first()
        eval_data = json.loads(eval_phase.data) if eval_phase and eval_phase.data else {}
        for k in ['actual_l1', 'actual_l2', 'actual_l3', 'actual_l4']:
            val = request.form.get(k, '').strip()
            if val:
                eval_data[k] = val
            elif k in eval_data:
                del eval_data[k]
        if eval_phase:
            eval_phase.data = json.dumps(eval_data)
        else:
            eval_phase = Phase(request_id=req_id, phase_name='evaluation', status='Not Started',
                data=json.dumps(eval_data), started_at=datetime.utcnow())
            db.session.add(eval_phase)
        db.session.commit()
        flash('Kirkpatrick actual results saved', 'success')
        return redirect(url_for('request_detail', req_id=req_id))

    @app.route('/requests/<int:req_id>/interview', methods=['POST'])
    @login_required
    def add_interview(req_id):
        iv = InterviewGuide(request_id=req_id, stakeholder_name=request.form['stakeholder_name'],
            stakeholder_role=request.form['stakeholder_role'],
            interview_date=datetime.strptime(request.form['interview_date'], '%Y-%m-%d').date(),
            questions=request.form['questions'], key_findings=request.form['key_findings'],
            submitted_by=current_user.id)
        db.session.add(iv)
        db.session.commit()
        flash('Interview recorded', 'success')
        return redirect(url_for('request_detail', req_id=req_id))

    @app.route('/requests/<int:req_id>/tags/add', methods=['POST'])
    @login_required
    def add_tag(req_id):
        tag = ContentTag(request_id=req_id, tag=request.form['tag'])
        db.session.add(tag)
        db.session.commit()
        flash('Tag added', 'success')
        return redirect(url_for('request_detail', req_id=req_id))

    @app.route('/requests/<int:req_id>/kirkpatrick/add', methods=['POST'])
    @login_required
    def add_kirkpatrick(req_id):
        kr = KirkpatrickResult(request_id=req_id, level=int(request.form['level']),
            metric_name=request.form['metric_name'],
            metric_value=float(request.form['metric_value'] or 0),
            target_value=float(request.form['target_value'] or 0),
            data_source=request.form['data_source'], assessed_at=datetime.utcnow())
        db.session.add(kr)
        if int(request.form['level']) == 4:
            req = Request.query.get_or_404(req_id)
            req.actual_roi = float(request.form['metric_value'] or 0)
        db.session.commit()
        flash('Kirkpatrick result recorded', 'success')
        return redirect(url_for('request_detail', req_id=req_id))

    @app.route('/analytics/export/<fmt>')
    @login_required
    def export_all_analytics(fmt):
        from flask import make_response
        import csv as csv_mod, io as io_mod
        from openpyxl import Workbook
        from openpyxl.styles import Font as XlFont, PatternFill as XlFill
        all_stats = LearningStatistic.query.all()
        if fmt == 'csv':
            output = io_mod.StringIO()
            w = csv_mod.writer(output)
            w.writerow(['Intervention','Learner Name','Email','Department','Status','Completion %','Score','Passed','Time Spent (min)','Attempts','Cohort'])
            for s in all_stats:
                req_title = s.request.title if s.request else ''
                cohort_name = s.cohort.name if s.cohort else ''
                w.writerow([req_title, s.learner_name or '', s.learner_email or '', s.department or '', s.completion_status or '',
                    s.completion_pct or '', s.score or '', 'Yes' if s.passed else 'No' if s.passed is not None else '', s.time_spent_minutes or '', s.assessment_attempts or '', cohort_name])
            response = make_response(output.getvalue())
            response.headers['Content-Disposition'] = 'attachment; filename=all_learning_analytics.csv'
            response.headers['Content-type'] = 'text/csv'
            return response
        # Excel export
        wb = Workbook(); ws = wb.active; ws.title = 'All Analytics'
        headers = ['Intervention','Learner Name','Email','Department','Status','Completion %','Score','Passed','Time Spent (min)','Attempts','Cohort']
        hdr_fill = XlFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = XlFont(bold=True, color='FFFFFF', size=10); cell.fill = hdr_fill
        for i, s in enumerate(all_stats, 2):
            req_title = s.request.title if s.request else ''
            cohort_name = s.cohort.name if s.cohort else ''
            row_data = [req_title, s.learner_name or '', s.learner_email or '', s.department or '', s.completion_status or '',
                s.completion_pct or '', s.score or '', 'Yes' if s.passed else 'No' if s.passed is not None else '', s.time_spent_minutes or '', s.assessment_attempts or '', cohort_name]
            for c, v in enumerate(row_data, 1):
                ws.cell(row=i, column=c, value=v).font = XlFont(size=9)
        for i, w in enumerate([30,25,30,20,15,12,10,10,15,10,20], 1):
            ws.column_dimensions[chr(64+i)].width = w
        output = io_mod.BytesIO()
        wb.save(output); output.seek(0)
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = 'attachment; filename=all_learning_analytics.xlsx'
        response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response

    @app.route('/analytics')
    @login_required
    def analytics():
        # Exclude Rejected requests from all counts
        active = Request.query.filter(Request.review_status != 'Rejected')
        total = active.count()
        by_status = db.session.query(Request.status, db.func.count(Request.id)).filter(Request.review_status != 'Rejected').group_by(Request.status).all()
        by_priority = db.session.query(Request.priority, db.func.count(Request.id)).filter(Request.review_status != 'Rejected').group_by(Request.priority).all()
        by_phase = db.session.query(Request.current_phase, db.func.count(Request.id)).filter(Request.review_status != 'Rejected').group_by(Request.current_phase).all()

        # Learning statistics overview
        all_stats = LearningStatistic.query.all()
        total_learners = len(all_stats)
        avg_completion = round(sum(s.completion_pct or 0 for s in all_stats) / total_learners, 1) if total_learners > 0 else 0
        avg_score_all = round(sum(s.score or 0 for s in all_stats) / total_learners, 1) if total_learners > 0 else 0
        passed_count = sum(1 for s in all_stats if s.passed)
        pass_rate = round(passed_count / total_learners * 100, 1) if total_learners > 0 else 0

        # Per-request stats
        requests_with_stats = []
        for r in Request.query.filter(Request.review_status != 'Rejected').all():
            stats = r.learning_stats.all()
            if stats:
                requests_with_stats.append({
                    'id': r.id, 'title': r.title,
                    'count': len(stats),
                    'avg_pct': r.avg_completion_pct(),
                    'avg_score': r.avg_score(),
                    'pass_rate': r.pass_rate()
                })

        completions = Phase.query.filter(Phase.status=='Completed', Phase.completed_at.isnot(None), Phase.started_at.isnot(None)).all()
        avg_times = {}
        for c in completions:
            dur = (c.completed_at - c.started_at).days
            if c.phase_name not in avg_times: avg_times[c.phase_name] = []
            avg_times[c.phase_name].append(dur)
        avg_phase_days = {p: round(sum(d)/len(d), 1) for p, d in avg_times.items()} if avg_times else {}

        kirk = db.session.query(KirkpatrickResult.level,
            db.func.avg(KirkpatrickResult.metric_value).label('avg'),
            db.func.avg(KirkpatrickResult.target_value).label('target'),
            db.func.count(KirkpatrickResult.id).label('count')
        ).group_by(KirkpatrickResult.level).all()

        # Kirpatrick: Projected vs Actual from all interventions
        projected_kirk = {'l1': '', 'l2': '', 'l3': '', 'l4': ''}
        actual_kirk = {'l1': '', 'l2': '', 'l3': '', 'l4': ''}
        analysis_phases = Phase.query.filter_by(phase_name='analysis').all()
        eval_phases = Phase.query.filter_by(phase_name='evaluation').all()
        
        # Collect projected from all analysis phases
        for p in analysis_phases:
            d = json.loads(p.data) if p.data else {}
            for k in ['projected_l1', 'projected_l2', 'projected_l3', 'projected_l4']:
                if d.get(k):
                    short = k.replace('projected_', '')
                    if projected_kirk[short]:
                        projected_kirk[short] += ' / '
                    projected_kirk[short] += str(d[k])
        
        # Collect actual from all evaluation phases
        for p in eval_phases:
            d = json.loads(p.data) if p.data else {}
            for k in ['actual_l1', 'actual_l2', 'actual_l3', 'actual_l4']:
                if d.get(k):
                    short = k.replace('actual_', '')
                    if actual_kirk[short]:
                        actual_kirk[short] += ' / '
                    actual_kirk[short] += str(d[k])

        # For L2: use manually entered actual if available, else fall back to computed avg from stats
        # Per-request L2 actual from uploaded stats
        per_request_l2 = {}
        for r in Request.query.all():
            stats = r.learning_stats.all()
            scores = [s.score for s in stats if s.score is not None]
            if scores:
                per_request_l2[r.id] = round(sum(scores) / len(scores), 1)
        
        # Compute overall L2 from stats (for interventions that don't have manual L2 entered)
        all_scores = [s.score for s in all_stats if s.score is not None]
        actual_l2_from_stats = round(sum(all_scores) / len(all_scores), 1) if all_scores else None
        # Only use computed L2 if no manual entries exist
        computed_only = [s for s in all_scores]  # all scores
        if not actual_kirk['l2'] and actual_l2_from_stats:
            actual_kirk['l2'] = f'{actual_l2_from_stats}% (from {len(all_scores)} learners)'

        blockers = []
        for p in Phase.query.filter(Phase.status=='Blocked').all():
            data = json.loads(p.data) if p.data else {}
            req = Request.query.get(p.request_id)
            blockers.append({'id': p.request_id, 'title': req.title if req else 'Unknown',
                'phase': p.phase_name, 'blocker': data.get('blocker', 'Unknown'), 'since': p.updated_at})

        escalations = Request.query.filter(Request.escalation_level>0).order_by(Request.escalation_level.desc()).all()
        roi_data = Request.query.filter(Request.projected_roi>0).all()

        return render_template('analytics.html', total=total, by_status=by_status,
            by_priority=by_priority, by_phase=by_phase, avg_phase_days=avg_phase_days,
            kirk=kirk, blockers=blockers, escalations=escalations, roi_data=roi_data,
            total_learners=total_learners, avg_completion=avg_completion, avg_score_all=avg_score_all,
            pass_rate=pass_rate, requests_with_stats=requests_with_stats,
            projected_kirk=projected_kirk, actual_kirk=actual_kirk, per_request_l2=per_request_l2)

    @app.route('/requests/<int:req_id>/analytics')
    @login_required
    def request_analytics(req_id):
        req = Request.query.get_or_404(req_id)
        stats = req.learning_stats.all()
        return render_template('request_analytics.html', req=req, stats=stats)

    @app.route('/admin')
    @login_required
    def admin():
        if not current_user.is_learning_expert():
            flash('Access denied', 'danger')
            return redirect(url_for('dashboard'))
        users = User.query.all()
        return render_template('admin.html', users=users)

    @app.route('/admin/users/add', methods=['POST'])
    @login_required
    def add_user():
        if not current_user.is_learning_expert():
            return redirect(url_for('dashboard'))
        user = User(email=request.form['email'], name=request.form['name'],
            role=request.form['role'], title=request.form['title'])
        user.set_password(request.form['password'])
        db.session.add(user)
        db.session.commit()
        flash('User added', 'success')
        return redirect(url_for('admin'))

    @app.route('/admin/users/<int:user_id>/edit', methods=['POST'])
    @login_required
    def edit_user(user_id):
        if not current_user.is_learning_expert():
            return redirect(url_for('dashboard'))
        user = User.query.get_or_404(user_id)
        user.name = request.form['name']
        user.email = request.form['email']
        user.role = request.form['role']
        user.title = request.form['title']
        if request.form.get('password'):
            user.set_password(request.form['password'])
        db.session.commit()
        flash('User updated', 'success')
        return redirect(url_for('admin'))

    @app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
    @login_required
    def delete_user(user_id):
        if not current_user.is_learning_expert():
            return redirect(url_for('dashboard'))
        if current_user.id == user_id:
            flash('Cannot delete yourself', 'danger')
            return redirect(url_for('admin'))
        user = User.query.get_or_404(user_id)
        db.session.delete(user)
        db.session.commit()
        flash('User deleted', 'success')
        return redirect(url_for('admin'))

    @app.route('/admin/users/upload-bulk', methods=['POST'])
    @login_required
    def bulk_upload_users():
        if not current_user.is_learning_expert():
            return redirect(url_for('dashboard'))
        if request.form.get('download_template'):
            import csv as csv_mod
            from flask import make_response
            output = io.StringIO()
            w = csv_mod.writer(output)
            w.writerow(['name','email','password','role','title'])
            w.writerow(['John Doe','john@addie.app','password123','learning_expert','Learning Designer'])
            w.writerow(['Jane Smith','jane@addie.app','password123','manager','Learning Manager'])
            response = make_response(output.getvalue())
            response.headers['Content-Disposition'] = 'attachment; filename=users_bulk_template.csv'
            response.headers['Content-type'] = 'text/csv'
            return response
        file = request.files.get('file')
        if not file or file.filename == '':
            flash('No file selected', 'danger')
            return redirect(url_for('admin'))
        import csv as csv_mod, io as io_mod
        content = file.read().decode('utf-8-sig').strip()
        first_line = content.split('\n')[0] if content else ''
        delimiter = ';' if ';' in first_line and first_line.count(';') >= first_line.count(',') else ','
        reader = csv_mod.DictReader(io_mod.StringIO(content), delimiter=delimiter)
        imported = 0; errors = 0
        for row in reader:
            try:
                email = row.get('email', '').strip().lower()
                name = row.get('name', '').strip()
                password = row.get('password', '').strip()
                role = row.get('role', 'manager').strip().lower()
                title = row.get('title', '').strip()
                if not email or not name:
                    errors += 1
                    continue
                if User.query.filter_by(email=email).first():
                    errors += 1
                    continue
                user = User(email=email, name=name, role=role, title=title)
                user.set_password(password if password else 'changeme123')
                db.session.add(user)
                imported += 1
            except Exception as e:
                errors += 1
        db.session.commit()
        flash(f'{imported} users imported, {errors} skipped (duplicates or missing data)', 'success')
        return redirect(url_for('admin'))

    # ── LNA BUDGET CYCLES ──
    @app.route('/lna-cycles')
    @login_required
    def lna_cycles():
        cycles = LNACycle.query.order_by(LNACycle.start_date.desc()).all()
        return render_template('lna_cycles.html', cycles=cycles)

    @app.route('/lna-cycles/new', methods=['POST'])
    @login_required
    def new_lna_cycle():
        budget = request.form.get('budget_amount', '0')
        c = LNACycle(name=request.form['name'], description=request.form.get('description',''),
            budget_amount=float(budget) if budget else 0,
            start_date=datetime.strptime(request.form['start_date'],'%Y-%m-%d').date(),
            end_date=datetime.strptime(request.form['end_date'],'%Y-%m-%d').date(),
            status=request.form.get('status','Planning'))
        db.session.add(c); db.session.commit()
        flash('LNA cycle created','success'); return redirect(url_for('lna_cycles'))

    @app.route('/lna-cycles/<int:cid>')
    @login_required
    def lna_cycle_detail(cid):
        cycle = LNACycle.query.get_or_404(cid)
        departments = cycle.department_budgets.order_by(DepartmentBudget.department_name).all()
        interventions = cycle.requests.order_by(Request.created_at.desc()).all()

        # Totals
        total_allocated = sum(d.allocated_budget for d in departments)
        total_spent_dept = sum(d.spent_budget for d in departments)
        total_cost_interventions = sum(r.allocated_cost or 0 for r in interventions)
        total_budget = cycle.budget_amount

        return render_template('lna_cycle_detail.html', cycle=cycle, departments=departments,
            interventions=interventions, total_allocated=total_allocated,
            total_spent_dept=total_spent_dept, total_cost_interventions=total_cost_interventions,
            total_budget=total_budget)

    @app.route('/lna-cycles/<int:cid>/departments/add', methods=['POST'])
    @login_required
    def add_department_budget(cid):
        db = DepartmentBudget(lna_cycle_id=cid,
            department_name=request.form['department_name'],
            allocated_budget=float(request.form.get('allocated_budget') or 0),
            spent_budget=float(request.form.get('spent_budget') or 0),
            notes=request.form.get('notes',''))
        db.session.add(db); db.session.commit()
        flash('Department budget added','success'); return redirect(url_for('lna_cycle_detail', cid=cid))

    @app.route('/lna-cycles/dept/<int:did>/update', methods=['POST'])
    @login_required
    def update_dept_budget(did):
        db = DepartmentBudget.query.get_or_404(did)
        db.allocated_budget = float(request.form.get('allocated_budget') or db.allocated_budget)
        db.spent_budget = float(request.form.get('spent_budget') or db.spent_budget)
        db.notes = request.form.get('notes', db.notes)
        db.session.commit(); flash('Department budget updated','success')
        return redirect(url_for('lna_cycle_detail', cid=db.lna_cycle_id))

    @app.route('/lna-cycles/dept/<int:did>/delete', methods=['POST'])
    @login_required
    def delete_dept_budget(did):
        db = DepartmentBudget.query.get_or_404(did)
        cid = db.lna_cycle_id
        db.session.delete(db); db.session.commit()
        flash('Department removed','success'); return redirect(url_for('lna_cycle_detail', cid=cid))

    @app.route('/lna-cycles/<int:cid>/request/<int:rid>/cost', methods=['POST'])
    @login_required
    def update_request_cost(cid, rid):
        req = Request.query.get_or_404(rid)
        req.allocated_cost = float(request.form.get('allocated_cost') or 0)
        db.session.commit(); flash('Cost updated','success')
        return redirect(url_for('lna_cycle_detail', cid=cid))

    @app.route('/lna-cycles/<int:cid>/close', methods=['POST'])
    @login_required
    def close_lna_cycle(cid):
        c = LNACycle.query.get_or_404(cid); c.status = 'Closed'
        db.session.commit(); flash('Cycle closed','success'); return redirect(url_for('lna_cycles'))

    # ── CALENDAR VIEW ──
    @app.route('/calendar')
    @login_required
    def calendar_view():
        events = TrainingEvent.query.order_by(TrainingEvent.start_date).all()
        requests = Request.query.order_by(Request.created_at.desc()).all()
        cycles = LNACycle.query.filter_by(status='Planning').all()
        users = User.query.all()
        forecasts = ResourceForecast.query.all()

        # Also load phase activities as virtual events for calendar display
        activities = PhaseActivity.query.filter(PhaseActivity.request_id.isnot(None)).all()
        activity_events = []
        from collections import namedtuple
        ActivityEvent = namedtuple('ActivityEvent', ['id','title','start_date','end_date','colour','event_type','status','trainer','request_id','notes'])
        for a in activities:
            activity_events.append(ActivityEvent(
                id=a.id,
                title=f'📌 {a.activity_name[:30]}',
                start_date=a.start_date,
                end_date=a.end_date,
                colour='#6C3483',
                event_type='activity',
                status=a.status,
                trainer=a.owner,
                request_id=a.request_id,
                notes=f'{a.phase_name} phase · {a.owner.name if a.owner else ""}'
            ))
        from datetime import date
        today = date.today()
        cal_month = request.args.get('month', today.month)
        cal_year = request.args.get('year', today.year)
        try: cal_month = int(cal_month); cal_year = int(cal_year)
        except: cal_month = today.month; cal_year = today.year
        if cal_month < 1: cal_month = 12; cal_year -= 1
        if cal_month > 12: cal_month = 1; cal_year += 1
        import calendar
        cal_days = calendar.monthrange(cal_year, cal_month)[1]
        first_day_of_month = (calendar.weekday(cal_year, cal_month, 1) + 1) % 7
        prev_month = cal_month - 1 if cal_month > 1 else 12
        prev_year = cal_year if cal_month > 1 else cal_year - 1
        next_month = cal_month + 1 if cal_month < 12 else 1
        next_year = cal_year if cal_month < 12 else cal_year + 1
        month_names = {1:'January',2:'February',3:'March',4:'April',5:'May',6:'June',
            7:'July',8:'August',9:'September',10:'October',11:'November',12:'December'}
        # Merge phase activities into events (purple, shown on calendar)
        all_events = list(events) + activity_events
        return render_template('calendar.html', events=all_events, requests=requests, cycles=cycles, users=users,
            forecasts=forecasts, cal_month=cal_month, cal_year=cal_year, cal_days=cal_days,
            first_day_of_month=first_day_of_month,
            prev_month=prev_month, prev_year=prev_year, next_month=next_month, next_year=next_year,
            month_names=month_names, today_day=today.day, today_month=today.month, today_year=today.year)

    @app.route('/calendar/events/add', methods=['POST'])
    @login_required
    def add_training_event():
        req_id = request.form.get('request_id')
        e = TrainingEvent(
            request_id=int(req_id) if req_id and int(req_id) > 0 else None,
            title=request.form['title'],
            start_date=datetime.strptime(request.form['start_date'],'%Y-%m-%d').date(),
            end_date=datetime.strptime(request.form['end_date'],'%Y-%m-%d').date(),
            trainer_id=int(request.form['trainer_id']) if request.form.get('trainer_id') else None,
            event_type=request.form.get('event_type','planned'),
            colour=request.form.get('colour','#2E86AB'),
            status=request.form.get('status','Scheduled'),
            location=request.form.get('location',''),
            notes=request.form.get('notes',''))
        db.session.add(e); db.session.commit()
        log_audit('Added training event', 'TrainingEvent', e.id, f'{e.title}')
        flash('Training event added','success'); return redirect(url_for('calendar_view'))

    @app.route('/calendar/events/<int:eid>/delete', methods=['POST'])
    @login_required
    def delete_training_event(eid):
        db.session.delete(TrainingEvent.query.get_or_404(eid)); db.session.commit()
        flash('Event deleted','success'); return redirect(url_for('calendar_view'))

    # ── KANBAN / WORKLOAD VIEW ──
    @app.route('/kanban')
    @login_required
    def kanban_view():
        allocations = ResourceAllocation.query.all()
        users = User.query.all()
        inbox = Request.query.filter_by(current_phase='inbox').order_by(Request.created_at.desc()).all()
        analysis = Request.query.filter_by(current_phase='analysis').all()
        design = Request.query.filter_by(current_phase='design').all()
        development = Request.query.filter_by(current_phase='development').all()
        implementation = Request.query.filter_by(current_phase='implementation').all()
        evaluation = Request.query.filter_by(current_phase='evaluation').all()
        completed = Request.query.filter_by(status='Completed').all()
        return render_template('kanban.html', allocations=allocations, users=users,
            inbox=inbox, analysis=analysis, design=design, development=development,
            implementation=implementation, evaluation=evaluation, completed=completed)

    @app.route('/kanban/allocate', methods=['POST'])
    @login_required
    def allocate_resource():
        inferred = None
        if request.form.get('inferred_completion'):
            inferred = datetime.strptime(request.form['inferred_completion'],'%Y-%m-%d').date()
        a = ResourceAllocation(request_id=int(request.form['request_id']),
            user_id=int(request.form['user_id']),
            inferred_completion=inferred,
            phase_responsible=request.form.get('phase_responsible',''),
            status='Allocated')
        db.session.add(a)
        req = Request.query.get(int(request.form['request_id']))
        if req and request.form.get('phase_responsible'):
            req.assigned_to_id = int(request.form['user_id'])
        db.session.commit()
        flash('Resource allocated','success'); return redirect(url_for('kanban_view'))

    @app.route('/kanban/allocate/<int:aid>/complete', methods=['POST'])
    @login_required
    def complete_allocation(aid):
        a = ResourceAllocation.query.get_or_404(aid)
        a.status = 'Complete'
        a.actual_completion = datetime.utcnow().date()
        if request.form.get('notes'): a.notes = request.form['notes']
        db.session.commit(); flash('Allocation completed','success'); return redirect(url_for('kanban_view'))

    @app.route('/kanban/allocate/<int:aid>/update-completion', methods=['POST'])
    @login_required
    def update_completion(aid):
        a = ResourceAllocation.query.get_or_404(aid)
        if request.form.get('inferred_completion'):
            a.inferred_completion = datetime.strptime(request.form['inferred_completion'],'%Y-%m-%d').date()
        db.session.commit(); flash('Completion date updated','success'); return redirect(url_for('kanban_view'))

    # ── LEARNING STATISTICS UPLOAD / DOWNLOAD ──
    @app.route('/requests/<int:req_id>/stats/download-template')
    @login_required
    def download_stats_template(req_id):
        req = Request.query.get_or_404(req_id)
        import csv
        from flask import make_response
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['learner_name','learner_email','department','completion_status','completion_pct','score','time_spent_minutes','assessment_attempts','passed','completed_date'])
        writer.writerow(['John Doe','john@example.com','Operations','Completed',100,85,120,1,'TRUE','2026/06/15'])
        writer.writerow(['Jane Smith','jane@example.com','Service Desk','In Progress',60,70,45,0,'FALSE',''])
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename=learning_stats_template_{req_id}.csv'
        response.headers['Content-type'] = 'text/csv'
        return response

    @app.route('/requests/<int:req_id>/stats/upload', methods=['POST'])
    @login_required
    def upload_stats(req_id):
        req = Request.query.get_or_404(req_id)
        file = request.files.get('file')
        if not file or file.filename == '':
            flash('No file selected', 'danger'); return redirect(url_for('request_detail', req_id=req_id))

        import csv, io, re
        content = file.read().decode('utf-8-sig')
        content = content.strip()
        first_line = content.split('\n')[0] if content else ''
        delimiter = ';' if ';' in first_line and first_line.count(';') >= first_line.count(',') else ','
        reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
        batch = datetime.utcnow().strftime('%Y%m%d%H%M%S')
        row_count = 0
        headers = reader.fieldnames or []
        # Accept either 'completed_date' or no date column
        has_date = 'completed_date' in headers

        def clean_num(val):
            """Clean a value that may contain % signs, time formats, or be empty"""
            if not val: return None
            val = str(val).strip().replace('%', '').replace(' ', '')
            # Handle time format like 00:17:26 → convert to minutes
            if re.match(r'^\d{1,2}:\d{2}:\d{2}$', val):
                parts = val.split(':')
                return float(parts[0]) * 60 + float(parts[1]) + float(parts[2]) / 60
            if re.match(r'^\d{1,2}:\d{2}$', val):
                parts = val.split(':')
                return float(parts[0]) * 60 + float(parts[1])
            try:
                return float(val)
            except ValueError:
                return None

        for row in reader:
            try:
                comp_pct = clean_num(row.get('completion_pct', ''))
                scr = clean_num(row.get('score', ''))
                time_min = clean_num(row.get('time_spent_minutes', ''))
                attempts_raw = row.get('assessment_attempts', '').strip()
                attempts = int(attempts_raw) if attempts_raw and attempts_raw.replace('.','','').isdigit() else None
                stat = LearningStatistic(
                    request_id=req_id,
                    learner_name=row.get('learner_name',''),
                    learner_email=row.get('learner_email',''),
                    department=row.get('department',''),
                    completion_status=row.get('completion_status','Not Started'),
                    completion_pct=comp_pct if comp_pct is not None else 0,
                    score=scr,
                    time_spent_minutes=time_min,
                    assessment_attempts=attempts,
                    passed=row.get('passed','').upper() in ('TRUE','1','YES','Y'),
                    upload_batch=batch
                )
                db.session.add(stat)
                row_count += 1
            except Exception as e:
                flash(f'Error on row {row_count+2}: {str(e)}', 'warning')

        batch_record = LearningUploadBatch(request_id=req_id, file_name=file.filename,
            rows_imported=row_count, uploaded_by=current_user.id)
        db.session.add(batch_record)
        db.session.commit()
        flash(f'{row_count} learner records imported', 'success')
        return redirect(url_for('request_detail', req_id=req_id))

    @app.route('/stats/<int:stat_id>/edit', methods=['POST'])
    @login_required
    def edit_learning_stat(stat_id):
        stat = LearningStatistic.query.get_or_404(stat_id)
        req_id = stat.request_id
        stat.learner_name = request.form.get('learner_name', stat.learner_name)
        stat.learner_email = request.form.get('learner_email', stat.learner_email)
        stat.department = request.form.get('department', stat.department)
        stat.completion_status = request.form.get('completion_status', stat.completion_status)
        stat.completion_pct = float(request.form.get('completion_pct', stat.completion_pct or 0) or 0)
        stat.score = float(request.form.get('score', stat.score or 0) or 0) if request.form.get('score') else None
        stat.time_spent_minutes = float(request.form.get('time_spent_minutes', stat.time_spent_minutes or 0) or 0) if request.form.get('time_spent_minutes') else None
        stat.assessment_attempts = int(request.form.get('assessment_attempts', stat.assessment_attempts or 0) or 0) if request.form.get('assessment_attempts') else None
        stat.passed = request.form.get('passed', '').upper() in ('TRUE','1','YES','Y') if request.form.get('passed') else stat.passed
        db.session.commit()
        flash('Learner record updated', 'success')
        return redirect(url_for('request_detail', req_id=req_id))

    @app.route('/stats/<int:stat_id>/delete', methods=['POST'])
    @login_required
    def delete_learning_stat(stat_id):
        stat = LearningStatistic.query.get_or_404(stat_id)
        req_id = stat.request_id
        db.session.delete(stat)
        db.session.commit()
        flash('Learner record deleted', 'success')
        return redirect(url_for('request_detail', req_id=req_id))

    @app.route('/requests/<int:req_id>/stats/move-to-cohort', methods=['POST'])
    @login_required
    def move_stats_to_cohort(req_id):
        """Move multiple learner records from request-level to a cohort"""
        cohort_id = request.form.get('cohort_id', '').strip()
        stat_ids = request.form.getlist('stat_ids')
        if not cohort_id or not stat_ids:
            flash('Select a cohort and at least one learner record', 'warning')
            return redirect(url_for('request_detail', req_id=req_id))
        cohort = Cohort.query.get(int(cohort_id))
        if not cohort:
            flash('Cohort not found', 'danger')
            return redirect(url_for('request_detail', req_id=req_id))
        moved = 0
        for sid in stat_ids:
            stat = LearningStatistic.query.get(int(sid))
            if stat and stat.request_id == req_id:
                stat.cohort_id = int(cohort_id)
                moved += 1
        db.session.commit()
        log_audit('Moved stats to cohort', 'Cohort', int(cohort_id),
            f'{moved} learner records moved from request {req_id} to cohort "{cohort.name}"')
        flash(f'{moved} learner record(s) moved to "{cohort.name}"', 'success')
        return redirect(url_for('request_detail', req_id=req_id))

    # ── BULK REQUEST UPLOAD ──
    @app.route('/requests/upload-bulk', methods=['GET','POST'])
    @login_required
    def bulk_request_upload():
        if request.method == 'POST':
            if request.form.get('download_template'):
                import csv, io as io_module
                from flask import make_response
                output = io_module.StringIO()
                writer = csv.writer(output)
                writer.writerow(['title','description','priority','request_type','status','current_phase',
                    'requester_name','requester_department','business_need','expected_outcome',
                    'target_audience','estimated_learners','deadline','created_at',
                    'projected_roi','allocated_cost','is_adhoc','completed_phases'])
                writer.writerow(['POPI Compliance Q1 2026','Annual POPI refresher for all ops staff','High',
                    'planned','Completed','evaluation','John Smith','Compliance',
                    'Regulatory requirement for annual POPI training','100% compliance by all staff',
                    'All staff','150','2026-02-28','2026-01-15','75000','20000','FALSE',
                    'analysis,design,development,implementation,evaluation'])
                writer.writerow(['INN8 Platform Training - Ops','Platform navigation training for new ops hires',
                    'Medium','planned','Completed','evaluation','Jane Doe','Operations',
                    'New ops hires need platform access training','Reduced support tickets',
                    'Ops team','45','2026-03-15','2026-02-01','30000','12000','FALSE',
                    'analysis,design,development,implementation,evaluation'])
                response = make_response(output.getvalue())
                response.headers['Content-Disposition'] = 'attachment; filename=stanlib_bulk_requests_template.csv'
                response.headers['Content-type'] = 'text/csv'
                return response

            file = request.files.get('file')
            if not file or file.filename == '':
                flash('No file selected', 'danger'); return redirect(url_for('bulk_request_upload'))

            import csv, io
            content = file.read().decode('utf-8-sig')
            # Strip any leading whitespace/newlines from BOM issues
            content = content.strip()
            if not content:
                flash('File is empty', 'danger'); return redirect(url_for('bulk_request_upload'))

            # Auto-detect delimiter: semicolon or comma
            first_line = content.split('\n')[0] if content else ''
            delimiter = ';' if ';' in first_line and (first_line.count(';') >= first_line.count(',')) else ','

            reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
            row_count = 0
            headers = reader.fieldnames or []
            if not headers:
                flash('No column headers found in CSV. The first row must contain column names.', 'danger')
                return redirect(url_for('bulk_request_upload'))

            required = ['title']
            missing = [c for c in required if c not in headers]
            if missing:
                flash(f'Missing required columns: {", ".join(missing)}', 'danger')
                return redirect(url_for('bulk_request_upload'))

            for row in reader:
                try:
                    title = row.get('title', '').strip()
                    if not title:
                        continue  # skip rows without a title

                    # Dedup: skip if a request with the same title already exists
                    existing = Request.query.filter_by(title=title).first()
                    if existing:
                        print(f'Skipping duplicate: {title}')
                        continue

                    req = Request(
                        title=title,
                        description=row.get('description', ''),
                        priority=row.get('priority', 'Medium'),
                        request_type=row.get('request_type', 'planned'),
                        status=row.get('status', 'New'),
                        current_phase=row.get('current_phase', 'inbox'),
                        requester_name=row.get('requester_name', ''),
                        requester_department=row.get('requester_department', ''),
                        business_need=row.get('business_need', ''),
                        expected_outcome=row.get('expected_outcome', ''),
                        target_audience=row.get('target_audience', ''),
                        estimated_learners=int(row.get('estimated_learners', 0) or 0),
                        deadline=datetime.strptime(row['deadline'].replace('/','-'), '%Y-%m-%d').date() if row.get('deadline') else None,
                        projected_roi=float(row.get('projected_roi', 0) or 0),
                        allocated_cost=float(row.get('allocated_cost', 0) or 0),
                        is_adhoc=row.get('is_adhoc', '').upper() in ('TRUE','1','YES','Y'),
                        created_at=datetime.strptime(row['created_at'].replace('/','-'), '%Y-%m-%d') if row.get('created_at') else datetime.utcnow()
                    )
                    db.session.add(req)
                    db.session.flush()

                    # Mark ADDIE phases as Completed based on completed_phases column
                    raw = row.get('completed_phases', '')
                    completed = [p.strip().lower() for p in raw.replace('"','').split(',') if p.strip()]
                    phase_map = {'inbox':'inbox','analysis':'analysis','design':'design','development':'development',
                                 'implementation':'implementation','evaluation':'evaluation'}
                    for pname, pcode in phase_map.items():
                        if pname in completed or pcode in completed:
                            phase = Phase(request_id=req.id, phase_name=pname, status='Completed',
                                data='{}', started_at=datetime.utcnow(), completed_at=datetime.utcnow())
                            db.session.add(phase)

                    # If all phases complete, mark request completed
                    if 'evaluation' in completed or 'completed' in completed:
                        req.status = 'Completed'
                        req.current_phase = 'evaluation'

                    row_count += 1
                except Exception as e:
                    flash(f'Error on request row {row_count+2}: {str(e)}', 'warning')

            db.session.commit()
            flash(f'{row_count} requests imported successfully', 'success')
            return redirect(url_for('request_list'))

        return render_template('bulk_upload.html')

    # ── AUDIT LOG HELPER ──
    def log_audit(action, entity_type, entity_id, details=''):
        from flask import request as flask_req
        try:
            log = AuditLog(user_id=current_user.id, action=action,
                entity_type=entity_type, entity_id=entity_id, details=str(details)[:500])
            db.session.add(log)
            db.session.commit()
        except:
            pass  # audit should never break the app

    # ── COHORTS ──
    @app.route('/requests/<int:req_id>/cohorts')
    @login_required
    def cohorts_for_request(req_id):
        req = Request.query.get_or_404(req_id)
        cohorts = Cohort.query.filter_by(request_id=req_id).order_by(Cohort.created_at.desc()).all()
        return render_template('cohorts.html', req=req, cohorts=cohorts)

    @app.route('/requests/<int:req_id>/cohorts/new', methods=['POST'])
    @login_required
    def new_cohort(req_id):
        lc = request.form.get('learner_count', '0').strip()
        try: learner_count = int(lc) if lc else 0
        except ValueError: learner_count = 0
        c = Cohort(request_id=req_id, name=request.form['name'],
            description=request.form.get('description',''),
            learner_count=learner_count,
            start_date=datetime.strptime(request.form['start_date'],'%Y-%m-%d').date() if request.form.get('start_date') else None,
            end_date=datetime.strptime(request.form['end_date'],'%Y-%m-%d').date() if request.form.get('end_date') else None,
            status=request.form.get('status','Planned'))
        db.session.add(c); db.session.commit()
        log_audit('Created cohort', 'Cohort', c.id, f'{c.name} for request {req_id}')
        flash('Cohort created','success'); return redirect(url_for('cohorts_for_request', req_id=req_id))

    @app.route('/cohorts/<int:cid>/delete', methods=['POST'])
    @login_required
    def delete_cohort(cid):
        c = Cohort.query.get_or_404(cid); req_id = c.request_id
        db.session.delete(c); db.session.commit()
        flash('Cohort deleted','success'); return redirect(url_for('cohorts_for_request', req_id=req_id))

    @app.route('/cohorts/<int:cid>/edit', methods=['POST'])
    @login_required
    def edit_cohort(cid):
        """Edit cohort name and description"""
        c = Cohort.query.get_or_404(cid); req_id = c.request_id
        new_name = request.form.get('name', '').strip()
        if new_name:
            old_name = c.name
            c.name = new_name
            c.description = request.form.get('description', c.description or '')
            db.session.commit()
            log_audit('Edited cohort', 'Cohort', cid, f'"{old_name}" → "{new_name}"')
            flash(f'Cohort renamed to "{new_name}"', 'success')
        return redirect(url_for('cohorts_for_request', req_id=req_id))

    @app.route('/cohorts/<int:cid>/add-learner', methods=['POST'])
    @login_required
    def add_cohort_learner(cid):
        """Add a learner directly to a cohort (name only, creates a LearningStatistic record)"""
        cohort = Cohort.query.get_or_404(cid)
        learner_name = request.form.get('learner_name', '').strip()
        if learner_name:
            learner_email = request.form.get('learner_email', '').strip()
            stat = LearningStatistic(
                request_id=cohort.request_id,
                cohort_id=cid,
                learner_name=learner_name,
                learner_email=learner_email,
                completion_status='Not Started',
                upload_batch=f'manual_{datetime.utcnow().strftime("%Y%m%d%H%M%S")}'
            )
            db.session.add(stat)
            db.session.commit()
            log_audit('Added learner to cohort', 'Cohort', cid, f'{learner_name}')
            flash(f'{learner_name} added to cohort', 'success')
        return redirect(url_for('cohorts_for_request', req_id=cohort.request_id))

    @app.route('/cohorts/<int:cid>/stats/upload', methods=['POST'])
    @login_required
    def upload_cohort_stats(cid):
        cohort = Cohort.query.get_or_404(cid)
        file = request.files.get('file')
        if not file or file.filename == '':
            flash('No file selected','danger'); return redirect(url_for('cohorts_for_request', req_id=cohort.request_id))
        import csv, io as io_mod, re
        content = file.read().decode('utf-8-sig').strip()
        first_line = content.split('\n')[0] if content else ''
        delimiter = ';' if ';' in first_line and first_line.count(';') >= first_line.count(',') else ','
        reader = csv.DictReader(io_mod.StringIO(content), delimiter=delimiter)
        batch = datetime.utcnow().strftime('%Y%m%d%H%M%S'); row_count = 0

        def clean_num(val):
            if not val: return None
            val = str(val).strip().replace('%', '').replace(' ', '')
            if re.match(r'^\d{1,2}:\d{2}:\d{2}$', val):
                parts = val.split(':'); return float(parts[0])*60 + float(parts[1]) + float(parts[2])/60
            if re.match(r'^\d{1,2}:\d{2}$', val):
                parts = val.split(':'); return float(parts[0])*60 + float(parts[1])
            try: return float(val)
            except ValueError: return None

        for row in reader:
            try:
                comp_pct = clean_num(row.get('completion_pct',''))
                scr = clean_num(row.get('score',''))
                time_min = clean_num(row.get('time_spent_minutes',''))
                attempts_raw = row.get('assessment_attempts','').strip()
                attempts = int(attempts_raw) if attempts_raw and attempts_raw.replace('.','','').isdigit() else None
                stat = LearningStatistic(request_id=cohort.request_id, cohort_id=cid,
                    learner_name=row.get('learner_name',''), learner_email=row.get('learner_email',''),
                    department=row.get('department',''),
                    completion_status=row.get('completion_status','Not Started'),
                    completion_pct=comp_pct if comp_pct is not None else 0,
                    score=scr, time_spent_minutes=time_min, assessment_attempts=attempts,
                    passed=row.get('passed','').upper() in ('TRUE','1','YES','Y'), upload_batch=batch)
                db.session.add(stat); row_count += 1
            except Exception as e:
                flash(f'Error row {row_count+2}: {str(e)}','warning')
        batch_record = LearningUploadBatch(request_id=cohort.request_id, cohort_id=cid,
            file_name=file.filename, rows_imported=row_count, uploaded_by=current_user.id)
        db.session.add(batch_record); db.session.commit()
        log_audit('Uploaded cohort stats', 'Cohort', cid, f'{row_count} rows from {file.filename}')
        flash(f'{row_count} records imported','success')
        return redirect(url_for('cohorts_for_request', req_id=cohort.request_id))

    @app.route('/cohorts/<int:cid>/stats/download-template')
    @login_required
    def download_cohort_stats_template(cid):
        import csv
        from flask import make_response
        output = io.StringIO()
        w = csv.writer(output)
        w.writerow(['learner_name','learner_email','department','completion_status','completion_pct','score','time_spent_minutes','assessment_attempts','passed','completed_date'])
        w.writerow(['John Doe','john@example.com','Operations','Completed',100,85,120,1,'TRUE','2026/06/15'])
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename=cohort_stats_template_{cid}.csv'
        response.headers['Content-type'] = 'text/csv'
        return response

    # ── AUDIT LOG VIEW ──
    @app.route('/audit-log')
    @login_required
    def audit_log_view():
        logs = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(200).all()
        return render_template('audit_log.html', logs=logs)

    # ── RESOURCE FORECAST ──
    @app.route('/forecast')
    @login_required
    def forecast_view():
        users = User.query.all()
        forecasts = ResourceForecast.query.order_by(ResourceForecast.week_start.desc()).all()
        requests_all = Request.query.order_by(Request.title).all()
        return render_template('forecast.html', users=users, forecasts=forecasts, requests=requests_all)

    @app.route('/forecast/add', methods=['POST'])
    @login_required
    def add_forecast():
        f = ResourceForecast(user_id=int(request.form['user_id']),
            request_id=int(request.form['request_id']) if request.form.get('request_id') else None,
            week_start=datetime.strptime(request.form['week_start'],'%Y-%m-%d').date(),
            week_end=datetime.strptime(request.form['week_end'],'%Y-%m-%d').date(),
            availability_pct=float(request.form.get('availability_pct',100)),
            allocated_pct=float(request.form.get('allocated_pct',0)),
            status=request.form.get('status','Forecasted'),
            notes=request.form.get('notes',''))
        db.session.add(f); db.session.commit()
        log_audit('Added forecast', 'ResourceForecast', f.id,
            f'{f.user.name} — {f.availability_pct}% avail, {f.allocated_pct}% allocated')
        flash('Forecast added','success'); return redirect(url_for('forecast_view'))

    @app.route('/forecast/<int:fid>/delete', methods=['POST'])
    @login_required
    def delete_forecast(fid):
        f = ResourceForecast.query.get_or_404(fid)
        db.session.delete(f); db.session.commit()
        flash('Forecast deleted','success'); return redirect(url_for('forecast_view'))

    @app.route('/api/pending-approvals')
    @login_required
    def pending_approvals():
        approvals = Approval.query.filter_by(approver_id=current_user.id, status='Pending').all()
        return jsonify([{'id': a.id, 'request_id': a.request_id,
            'request_title': a.request.title, 'phase': a.phase_name,
            'created': a.created_at.strftime('%Y-%m-%d')} for a in approvals])
